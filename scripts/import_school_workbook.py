#!/usr/bin/env python3
"""Convert the Teacher Appreciation workbook into the app's seed records."""

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


STATE_ABBREVIATIONS = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID",
    "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS",
    "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
    "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
    "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY",
    "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK",
    "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC",
    "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT",
    "Vermont": "VT", "Virginia": "VA", "Washington": "WA", "West Virginia": "WV",
    "Wisconsin": "WI", "Wyoming": "WY",
}


def clean(value):
    if value is None:
        return ""
    value = str(value).strip()
    return "" if value.startswith("#") else value


def key(value):
    return re.sub(r"[^a-z0-9]", "", clean(value).lower())


def number(value):
    return int(value) if isinstance(value, (int, float)) and value >= 0 else 0


def first_email(value):
    match = re.search(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", clean(value), re.I)
    return match.group(0).lower() if match else ""


def split_location(location, fallback_city="", fallback_state=""):
    location = clean(location)
    if "," in location:
        city, state = [part.strip() for part in location.rsplit(",", 1)]
        return city, STATE_ABBREVIATIONS.get(state, state[:2].upper())
    if location:
        return location, ""
    return fallback_city, STATE_ABBREVIATIONS.get(fallback_state, fallback_state[:2].upper())


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input")
    parser.add_argument("output")
    args = parser.parse_args()

    workbook = load_workbook(args.input, data_only=True, read_only=True)

    directory = {}
    if "School directory" in workbook.sheetnames:
        for row in workbook["School directory"].iter_rows(min_row=2, values_only=True):
            name = clean(row[0] if len(row) > 0 else "")
            if not name:
                continue
            admin_indexes = [23, 27, 31, 35, 39, 41, 45, 49, 53]
            email_indexes = [25, 29, 33, 37, 43, 47, 51, 55]
            admin = next((clean(row[i]) for i in admin_indexes if i < len(row) and clean(row[i])), "")
            admin_email = next((first_email(row[i]) for i in email_indexes if i < len(row) and first_email(row[i])), "")
            directory[key(name)] = {
                "phone": clean(row[2] if len(row) > 2 else ""),
                "city": clean(row[3] if len(row) > 3 else ""),
                "state": clean(row[4] if len(row) > 4 else ""),
                "email": first_email(row[9] if len(row) > 9 else "") or admin_email,
                "admin": admin,
            }

    source = workbook["01 SCHOOLS"]
    colors = ["mint", "blue", "peach", "violet", "gold", "rose"]
    schools = []
    seen = set()

    for row in source.iter_rows(min_row=3, values_only=True):
        name = clean(row[0] if len(row) > 0 else "")
        code25 = clean(row[6] if len(row) > 6 else "")
        code24 = clean(row[9] if len(row) > 9 else "")
        if not name or (not code25 and not code24):
            continue
        identity = (key(name), code25.lower(), code24.lower())
        if identity in seen:
            continue
        seen.add(identity)

        match = directory.get(key(name), {})
        email = first_email(row[1] if len(row) > 1 else "") or match.get("email", "")
        phone = clean(row[12] if len(row) > 12 else "") or match.get("phone", "")
        city, state = split_location(
            row[13] if len(row) > 13 else "",
            match.get("city", ""),
            match.get("state", ""),
        )
        orders25 = number(row[11] if len(row) > 11 else 0)
        orders24 = number(row[10] if len(row) > 10 else 0)
        initials = "".join(part[0] for part in re.findall(r"[A-Za-z0-9]+", name)[:2]).upper() or "SC"
        school_id = len(schools) + 1
        schools.append({
            "id": school_id,
            "name": name,
            "schoolType": "regular",
            "district": "",
            "city": city,
            "state": state,
            "code": code25,
            "code2025": code25,
            "code2024": code24,
            "admin": match.get("admin", ""),
            "email": email,
            "phone": phone,
            "students": 0,
            "orders2026": 0,
            "orders2025": orders25,
            "orders2024": orders24,
            "status": "Not started",
            "progress": 0,
            "eligibility": "",
            "lastContact": "",
            "initials": initials[:2],
            "color": colors[(school_id - 1) % len(colors)],
        })

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(schools, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "schools": len(schools),
        "orders_2026": sum(s["orders2026"] for s in schools),
        "orders_2025": sum(s["orders2025"] for s in schools),
        "orders_2024": sum(s["orders2024"] for s in schools),
    }))


if __name__ == "__main__":
    main()
